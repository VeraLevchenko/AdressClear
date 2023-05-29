from openpyxl import load_workbook

from PyQt5 import QtCore, QtWidgets
import re

import func
import time

start = time.time() ## точка отсчета времени


# app = QtWidgets.QApplication([])
# wb_patch = QtWidgets.QFileDialog.getOpenFileName()[0]
# print(wb_patch)
# wb = load_workbook(wb_patch)

wb1 = load_workbook('D:/AdressClear/Src/rosreestr.xlsx')
table_input = wb1.active

wb2 = load_workbook("D:/AdressClear/Src/street_list.xlsx")
street_list = wb2.active

wb3 = load_workbook("D:/AdressClear/Src/snt_list.xlsx")
snt_list = wb3.active

wb4 = load_workbook("D:/AdressClear/Src/kooperativ_list.xlsx")
kooperativ_list = wb4.active

wb5 = load_workbook("D:/AdressClear/Src/kvartal.xlsx")
kvartal_list = wb5.active

def get_kooperativ(input_adress):
    kooperativ = "None"
    for j in range(1, kooperativ_list.max_row + 1):
        kooperativ = kooperativ_list[j][0].value
        index = input_adress.find(kooperativ)
        if index > 0:
            break
        else:
            kooperativ = "None"
    return kooperativ

def get_kvartal(input_adress):
    kvartal_clear = "None"
    for j in range(1, kvartal_list.max_row + 1):
        kvartal = kvartal_list[j][0].value
        kvartal_clear = kvartal_list[j][1].value
        index = input_adress.find(kvartal)
        if index > 0:
            break
        else:
            kvartal_clear = "None"
    return kvartal_clear

def get_snt(input_adress):
    snt = "None"
    for j in range(1, snt_list.max_row + 1):
        snt = snt_list[j][0].value
        index = input_adress.find(snt)
        if index > 0:
            break
        else:
            snt = "None"
    return snt

def get_street(input_adress):
    street = "None"
    for j in range(1, street_list.max_row + 1):
        street = street_list[j][0].value
        index = input_adress.find(street)
        if index > 0:
            break
        else:
            street = "None"
    return street


for i in range(2, table_input.max_row + 1):
    build = "None"
    korpus = "None"
    garag = "None"
    input_adress = str(table_input[i][18].value)
    print(input_adress)
    print(i)
    input_adress = input_adress.replace("-", '')
    input_adress = input_adress.replace(" ", '')
    input_adress = input_adress.replace("№", '')
    input_adress = input_adress.replace("Кемеровскаяобл", '')
    input_adress = input_adress.replace("обл.Кемеровская", '')
    input_adress = input_adress.replace("Кузнецкийрайон", '')
    input_adress = input_adress.replace("Орджоникидзевский", '')
    # print(input_adress)
    snt = get_snt(input_adress)
    # print("СНТ ", snt)
    kooperativ = get_kooperativ(input_adress)
    # print("Гаражный кооператив ", kooperativ)

    kvartal = get_kvartal(input_adress)
    # print("Квартал", kvartal)


    street = get_street(input_adress)
    # print("Улица: ", street)
    if street != "None":
        build = func.getBuild(input_adress, street)
        # print("Здание:", build)
        if build != "None":
            korpus = func.getKorpus(input_adress)
            # print("Корпус:", korpus)
    garag = func.getGarage(input_adress)
    # print("Гараж:", garag)

    table_input[i][0].value = snt
    table_input[i][1].value = kooperativ
    table_input[i][2].value = kvartal
    table_input[i][3].value = street
    table_input[i][4].value = build
    table_input[i][5].value = korpus
    table_input[i][6].value = garag

wb1.save('D:/AdressClear/Src/rosreestr.xlsx')
end = time.time() - start ## собственно время работы программы

print(end) ## вывод времени
