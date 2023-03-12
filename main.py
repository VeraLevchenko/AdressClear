from openpyxl import load_workbook
from PyQt5 import QtCore, QtWidgets
import re

app = QtWidgets.QApplication([])
wb_patch = QtWidgets.QFileDialog.getOpenFileName()[0]
print(wb_patch)
wb = load_workbook(wb_patch)
# wb = load_workbook('D:/project_Python/Adress_clear/source.xlsx')
sheet = wb.active

reestr_ul = load_workbook("D:/project_Python/Adress_clear/реестр_улиц.xlsx")
sheet_reestr_ul = reestr_ul.active

reestr_snt = load_workbook("D:/project_Python/Adress_clear/реестр_snt.xlsx")
sheet_reestr_snt = reestr_snt.active

# print(sheet_reestr_ul.max_row)
# ------------------------Поиск названия улицы----------------------------------------------------


def get_ul(_adres, _snt):
    _ul = ''
    if not _snt:
        for j in range(400, 452):
            print(sheet_reestr_ul[j][0].value)
            _ul = sheet_reestr_ul[j][0].value
            print(_ul)
            index = _adres.find(_ul)
            if index == -1:
                _ul = "Улица не найдена"
    else:
        _ul = _snt
    return _ul


def get_adres_without_ul(adres, ul):
    adres_without_ul = adres.replace(ul, '')
    return adres_without_ul


def get_number(adres):
    numbers = re.findall(r'\d+[-]\w|\d+', adres)
    return numbers


def get_snt(adres):
    snt = ''
    for j in range(1, sheet_reestr_snt.max_row):
        snt = sheet_reestr_snt[j][0].value
        index = adres.find(snt)
        if index == -1:
            snt = ""
    return snt


# перебирает адреса
for i in range(1, 43):
    adres = str(sheet[i][2].value)
    snt = get_snt(adres)
    ul = get_ul(adres, snt)
    sheet[i][3].value = ul
    adres_without_ul = get_adres_without_ul(adres, ul)
    numbers = get_number(adres_without_ul)
    k = 4
    # -----записывает числа из адреса в столбцы
    for number in numbers:
        sheet[i][k].value = number
        k += 1
wb.save('D:/project_Python/Adress_clear/source.xlsx')
# app.exec()
